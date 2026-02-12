"""
csv_excel_to_json.py — Convert CSV/Excel test cases to structured JSON.
Reads INPUT_DIR from agent_analyzer/.env to find the input file.
Usage: python csv_excel_to_json.py
"""

import csv, json, os, re, sys
from collections import OrderedDict

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── Load .env ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
ENV_PATH = os.path.join(REPO_ROOT, ".env")


def load_env(path):
    env = {}
    if not os.path.isfile(path):
        return env
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line and "=" in line and not line.startswith("#"):
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip().strip("'\"")
    return env


env = load_env(ENV_PATH)
INPUT_DIR = env.get("INPUT_DIR", "").strip()
INTERMEDIATE_DIR = env.get("INTERMEDIATE_DIR", "").strip()

# Resolve relative paths against repo root
if INPUT_DIR and not os.path.isabs(INPUT_DIR):
    INPUT_DIR = os.path.join(REPO_ROOT, INPUT_DIR)
if INTERMEDIATE_DIR and not os.path.isabs(INTERMEDIATE_DIR):
    INTERMEDIATE_DIR = os.path.join(REPO_ROOT, INTERMEDIATE_DIR)

if not INPUT_DIR or not os.path.isdir(INPUT_DIR):
    print(f"❌ INPUT_DIR not set or invalid in {ENV_PATH}")
    sys.exit(1)

# ... (rest of the file remains the same)




def sanitize(val):
    if val is None: return None
    s = str(val).strip()
    return None if s.lower() in ("none", "", "null", "n/a") else s


def find_input_file(directory):
    """Find the first .csv or .xlsx file in the directory."""
    for f in os.listdir(directory):
        if f.endswith((".csv", ".xlsx", ".xls")):
            return os.path.join(directory, f)
    return None


def read_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        with open(filepath, "r", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    elif ext in (".xlsx", ".xls"):
        if not openpyxl:
            print("❌ pip install openpyxl"); sys.exit(1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Test_Cases"] if "Test_Cases" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def convert(rows):
    groups = OrderedDict()
    for row in rows:
        tc_id = sanitize(row.get("TC_ID"))
        if tc_id:
            groups.setdefault(tc_id, []).append(row)

    modules = OrderedDict()
    for tc_id, tc_rows in groups.items():
        first = tc_rows[0]
        module = sanitize(first.get("Module")) or "Uncategorized"

        steps = []
        for r in tc_rows:
            step_no = int(sanitize(r.get("Step_No")) or len(steps) + 1)
            action = sanitize(r.get("Action")) or ""
            method = f"step_{step_no:02d}_{'_'.join(re.sub(r'[^a-zA-Z0-9 ]', '', action).lower().split())}"

            steps.append({
                "step_no": step_no,
                "action": action,
                "element_type": sanitize(r.get("Element_Type")) or "",
                "original_hint": sanitize(r.get("Element_Identifier_Hint")) or "",
                "resolved_locators": {
                    "value": "",
                    "confidence": 0.0,
                    "xpath": "",
                    "fallbacks": []
                },
                "input_data": sanitize(r.get("Input_Data")),
                "expected_result": sanitize(r.get("Expected_Result")) or "",
                "assertion": {
                    "type": sanitize(r.get("Assertion_Type")) or "",
                    "target": None,
                    "value": None
                },
                "method_name": method
            })

        tc = {
            "tc_id": tc_id,
            "title": sanitize(first.get("Test_Case_Title")) or "",
            "description": sanitize(first.get("Test_Case_Description")) or "",
            "priority": sanitize(first.get("Priority")) or "",
            "test_type": sanitize(first.get("Test_Type")) or "",
            "preconditions": sanitize(first.get("Preconditions")) or "",
            "steps": steps
        }
        modules.setdefault(module, []).append(tc)

    return [{"module": m, "test_cases": tcs} for m, tcs in modules.items()]


if __name__ == "__main__":
    input_file = find_input_file(INPUT_DIR)
    if not input_file:
        print(f"❌ No .csv or .xlsx file found in INPUT_DIR: {INPUT_DIR}")
        sys.exit(1)

    # Use INTERMEDIATE_DIR for output, creating it if needed
    if INTERMEDIATE_DIR and not os.path.exists(INTERMEDIATE_DIR):
        os.makedirs(INTERMEDIATE_DIR, exist_ok=True)
    
    output_dir = INTERMEDIATE_DIR if INTERMEDIATE_DIR and os.path.isdir(INTERMEDIATE_DIR) else INPUT_DIR
    output_file = os.path.join(output_dir, os.path.splitext(os.path.basename(input_file))[0] + "_output.json")

    print(f"📂 Input  : {input_file}")
    rows = read_file(input_file)
    result = convert(rows)

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    total_tc = sum(len(m["test_cases"]) for m in result)
    total_steps = sum(len(tc["steps"]) for m in result for tc in m["test_cases"])
    print(f"✅ Done — {len(result)} modules, {total_tc} test cases, {total_steps} steps")
    print(f"💾 Output : {output_file}")
